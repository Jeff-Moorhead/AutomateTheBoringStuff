#! python3

import openpyxl
import sys
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active

col = 1
for file_name in sys.argv[1:]:
    file = open(file_name, 'r')
    lines = file.readlines()
    row = 1
    max_width = 0

    for line in lines:
        sheet.cell(row=row, column=col).value = line
        col_letter = get_column_letter(col)

        if len(line) > max_width:
            max_width = len(line)
        row += 1
    
    sheet.column_dimensions[col_letter].width = max_width
    col += 1
    continue

wb.save('textSpreadsheet.xlsx')
