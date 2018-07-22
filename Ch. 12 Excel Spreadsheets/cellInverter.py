#! python3

import openpyxl
import sys
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(sys.argv[1])
new_workbook = openpyxl.Workbook()
sheet = wb.active
new_sheet = new_workbook.active
data = []
last_cell = get_column_letter(sheet.max_column) + str(sheet.max_row)

for row in sheet['A1':last_cell]:
    current_row = []
    for cell in row:
        current_row.append(cell.value)
    data.append(current_row)

for i in range(len(data)):
    for j in range(len(data[i])):
        new_sheet.cell(row=j+1, column=i+1).value = data[i][j]

new_workbook.save('invertedCells.xlsx')
