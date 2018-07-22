#! python3

"""Read census data from censuspopdata.xlsx and print the
total population and the number of census tracts for each
county."""

import openpyxl
import pprint

wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
population = 0
tracts = 0
current_county = sheet['C2'].value
counties = {}
col = list(sheet.columns)[2][1:]

for cell in col:
    index = current_county + ", " + sheet.cell(row=cell.row, column=2).value
    if cell.value == current_county:
        tracts += 1
        population += sheet.cell(row=cell.row, column=4).value
    else:
        if not index in counties.keys():
            counties[index] = {'Tracts': tracts, 'Population': population}
        current_county = cell.value
        population = sheet.cell(row=cell.row, column=4).value
        tracts = 1
        continue

with open('census.py', 'w') as file:
    file.write('census_data = ' + pprint.pformat(counties))
    file.close()

print("Census data stored successfully.")
