#! python3

import openpyxl
import sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active
dimensions = int(sys.argv[1])
max_row = dimensions + 1
max_col = get_column_letter(dimensions + 1)
last_cell = max_col + str(max_row)

sheet.title = f'Multiplication Table'

# Enter first row and column with labels in boldface font
for i in range(2, dimensions + 2):
    row_cell = sheet.cell(row=i, column=1)
    col_cell = sheet.cell(row=1, column=i)
    row_cell.value = col_cell.value = i - 1
    row_cell.font = col_cell.font = Font(bold=True)

# Populate the rest of the cells with products of labels
for row in sheet['B2':last_cell]:
    for cell in row:
        first_product = sheet['A' + str(cell.row)].value
        second_product = sheet[cell.column + str(1)].value
        cell.value = first_product * second_product

wb.save('multTables.xlsx')
