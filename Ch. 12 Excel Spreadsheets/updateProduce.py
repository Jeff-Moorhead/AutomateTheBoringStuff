#! python3

"""Searches the rows of a spreadsheet and updates the prices for
incorrectly priced items."""

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
price_updates = {'Celery': 1.19, 'Garlic': 3.07, 'Lemon': 1.27}
column = list(sheet.columns)[0]
counter = 0

for row in range(2, sheet.max_row + 1):
    product = sheet.cell(row=row, column=1).value
    if product in price_updates:
        sheet.cell(row=row, column=2).value = price_updates[product]
        counter += 1

print(f'{counter} items updated.')
wb.save('produceSalesUpdated.xlsx')
